from __future__ import annotations

import shutil
from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook
from sqlalchemy import select

from ..config import repo_root, ensure_data_dirs, default_db_path
from ..db import PricingSnapshot, session_scope


def _ensure_model_copies() -> List[Path]:
    root = repo_root()
    src_files = [
        root / "subscription_economics.xlsx",
        root / "subscriber_mix_model.xlsx",
    ]
    models_dir = root / "data" / "models"
    models_dir.mkdir(parents=True, exist_ok=True)
    copied: List[Path] = []
    for src in src_files:
        if not src.exists():
            continue
        dst = models_dir / src.name
        if not dst.exists():
            shutil.copy(src, dst)
        copied.append(dst)
    return copied


def _latest_pricing(db_path) -> List[Dict]:
    results: Dict[str, Dict] = {}
    with session_scope(db_path) as session:
        rows = (
            session.execute(
                select(PricingSnapshot)
                .where(PricingSnapshot.features.is_not(None))
                .order_by(PricingSnapshot.captured_at.desc())
            )
            .scalars()
            .all()
        )
        for snap in rows:
            if snap.company_id in results:
                continue
            data = snap.features or {}
            results[snap.company_id] = {
                "captured_at": snap.captured_at,
                "url": snap.url,
                "pricing": data.get("pricing", []),
            }
    out = []
    for cid, data in results.items():
        for entry in data["pricing"]:
            out.append(
                {
                    "company": cid,
                    "tier": entry.get("tier"),
                    "price_monthly": entry.get("price_monthly"),
                    "captured_at": data["captured_at"],
                    "source_url": data["url"],
                }
            )
    return out


def update_models(db_path=None):
    ensure_data_dirs()
    db_path = db_path or default_db_path()
    pricing_rows = _latest_pricing(db_path)
    model_paths = _ensure_model_copies()
    if not model_paths:
        raise FileNotFoundError("Source XLSX files not found in repo root.")

    for path in model_paths:
        wb = load_workbook(path)
        if "LatestPricing" in wb.sheetnames:
            ws = wb["LatestPricing"]
            wb.remove(ws)
        ws = wb.create_sheet("LatestPricing")
        headers = ["Company", "Tier", "PriceMonthly", "CapturedAt", "SourceURL"]
        ws.append(headers)
        for row in pricing_rows:
            ws.append(
                [
                    row.get("company"),
                    row.get("tier"),
                    row.get("price_monthly"),
                    row.get("captured_at"),
                    row.get("source_url"),
                ]
            )
        wb.save(path)
